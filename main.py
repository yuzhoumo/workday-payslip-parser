import os
import csv
import json
import argparse
import warnings
from datetime import datetime
from openpyxl import load_workbook

from io import TextIOWrapper
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any
from enum import Enum


class PayslipSection:
    SINGLE = 'SINGLE'
    GRID = 'GRID'
    SECTIONS = {
        'Company Information': SINGLE,
        'Payslip Information': SINGLE,
        'Current and YTD Totals': GRID,
        'Earnings': GRID,
        'Employee Taxes': GRID,
        'Pre Tax Deductions': GRID,
        'Post Tax Deductions': GRID,
        'Employer Paid Benefits': GRID,
        'Taxable Wages': GRID,
        'Withholding': GRID,
        'Payment Information': GRID,
    }


class OutputFormat(Enum):
    JSON = 'json'
    CSV = 'csv'


def parse_table_single_row(ws: Worksheet, start_row: Any, output: dict,
                           fmt: OutputFormat = OutputFormat.JSON):
    """
    Parses a single-row style table from the worksheet.

    The format is expected to be:
    - Title in the first column at `start_row`
    - Labels in the row `start_row + 1`
    - Values in the row `start_row + 2`

    Each key is constructed as: "<Title> - <Label>"

    Args:
        ws (Worksheet): The worksheet to parse.
        start_row (Any): The row index where the table starts.
        output (dict): Dictionary to store parsed key-value pairs.
        fmt (OutputFormat): Parsing output format (either json or csv).
    """
    key_prefix = ws.cell(start_row, 1).value
    label_row = start_row + 1
    data_row = start_row + 2
    col = 1

    if fmt == OutputFormat.JSON:
        output[key_prefix] = {}

    while ws.cell(label_row, col).value is not None:
        key_suffix = ws.cell(label_row, col).value
        val = ws.cell(data_row, col).value
        col += 1

        if fmt == OutputFormat.JSON:
            output[key_prefix][key_suffix] = val
        else:
            key = f'{key_prefix} - {key_suffix}'
            output[key] = val


def parse_table_grid(ws: Worksheet, start_row: Any, output: dict,
                     fmt: OutputFormat = OutputFormat.JSON):

    """
    Parses a grid-style table from the worksheet.

    The format is expected to be:
    - Title in the first column at `start_row`
    - Column headers in row `start_row + 1`
    - Row headers in column A starting at `start_row + 2`
    - Data values in the grid formed by headers

    Each key is constructed as: "<Title> - <Row Header> - <Column Header>"

    Args:
        ws (Worksheet): The worksheet to parse.
        start_row (Any): The row index where the table starts.
        output (dict): Dictionary to store parsed key-value pairs.
        fmt (OutputFormat): Parsing output format (either json or csv).
    """
    key_prefix = ws.cell(start_row, 1).value
    label_row = start_row + 1
    data_row = start_row + 2

    if fmt == OutputFormat.JSON:
        output[key_prefix] = []

    last_row = ws.max_row
    while ws.cell(data_row, 1).value not in PayslipSection.SECTIONS and data_row <= last_row:
        col = 2
        key_suffix_1 = ws.cell(data_row, 1).value

        if fmt == OutputFormat.JSON:
            output[key_prefix].append({})

        while ws.cell(label_row, col).value is not None:
            key_suffix_2 = ws.cell(label_row, col).value
            val = ws.cell(data_row, col).value
            col += 1

            if fmt == OutputFormat.JSON:
                output[key_prefix][-1]['Description'] = key_suffix_1
                output[key_prefix][-1][key_suffix_2] = val
            else:
                key = f'{key_prefix} - {key_suffix_1} - {key_suffix_2}'
                output[key] = val
        data_row += 1


def parse_payslip(fmt: OutputFormat, filename: str) -> dict:
    """
    Parses a payslip Excel file and extracts relevant information
    based on section headers.

    Args:
        fmt (OutputFormat): Format of the output file (either json or csv).
        filename (str): Path to the Excel (.xlsx) file.

    Returns:
        dict: A dictionary with keys as constructed from table headers
              and values as extracted data.
    """
    ws = load_workbook(filename).active
    if not ws:
        raise Exception(f"failed to load workbook: {filename}")

    payslip_data = {}
    for cell in ws['A']:
        match PayslipSection.SECTIONS.get(cell.value):
            case PayslipSection.SINGLE:
                parse_table_single_row(ws, cell.row, payslip_data, fmt)
            case PayslipSection.GRID:
                parse_table_grid(ws, cell.row, payslip_data, fmt)

    return payslip_data


def convert_to_csv_file(dir: str, out: TextIOWrapper, quiet: bool = False):
    """
    Converts all .xlsx payslip files in the given directory to a CSV file.

    Args:
        dir (str): Path to the directory containing Excel files.
        out (TextIOWrapper): Output file handle to write the CSV data.
        quiet (bool): Whether to print progress messages.
    """
    keys, rows = [], []
    for filename in os.listdir(dir):
        if filename.endswith('.xlsx'):
            full_path = os.path.join(dir, filename)
            if not quiet:
                print(f'Parsing {filename}...')
            row = parse_payslip(OutputFormat.CSV, full_path)
            rows.append(row)
            keys.extend(row.keys())

    # Ensure unique keys and maintain order
    keys = list(dict.fromkeys(keys))

    # Sort rows by date
    date_fmt, default = '%m/%d/%Y', '1/1/0001'
    header = 'Payslip Information - Check Date'
    sort_key = lambda r: datetime.strptime(r.get(header, default), date_fmt).date()
    rows.sort(key=sort_key)

    writer = csv.DictWriter(out, fieldnames=keys)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)


def convert_to_json_file(dir: str, out: TextIOWrapper, quiet: bool = False):
    """
    Converts all .xlsx payslip files in the given directory to a CSV file.

    Args:
        dir (str): Path to the directory containing Excel files.
        out (TextIOWrapper): Output file handle to write the CSV data.
        quiet (bool): Whether to print progress messages.
    """
    objs = []
    for filename in os.listdir(dir):
        if filename.endswith('.xlsx'):
            full_path = os.path.join(dir, filename)
            if not quiet:
                print(f'Parsing {filename}...')
            row = parse_payslip(OutputFormat.JSON, full_path)
            objs.append(row)

    # Sort objects by date
    date_fmt, default = '%m/%d/%Y', '1/1/0001'
    sort_key = lambda r: datetime.strptime(
        r['Payslip Information'].get('Check Date', default), date_fmt,
    ).date()
    objs.sort(key=sort_key)

    out.write(json.dumps({ 'Payslips': objs }, indent=2))


def main():
    """
    Entry point of the script. Parses CLI arguments and runs the conversion.
    """
    parser = argparse.ArgumentParser(
        description='Convert Workday payslip Excel files to a single CSV file.'
    )
    parser.add_argument(
        '-i', '--input-dir', type=str, default='.',
        help='Directory containing .xlsx files (default: current directory)'
    )
    parser.add_argument(
        '-o', '--output-file', type=str, default='output',
        help='Output CSV file name (default: output)'
    )
    parser.add_argument(
        '-f', '--format', type=str,
        choices=[e.value for e in OutputFormat],
        default=str(OutputFormat.JSON.value),
        help='Output format (default: json)'
    )
    parser.add_argument(
        '-q', '--quiet', action='store_true',
        help='Do not print logs while parsing'
    )

    args = parser.parse_args()

    # ignore only the “no default style” warning from openpyxl
    warnings.filterwarnings(
        "ignore",
        message="Workbook contains no default style, apply openpyxl's default"
    )

    match OutputFormat(args.format):
        case OutputFormat.JSON:
            with open(args.output_file + '.json', 'w', newline='') as f:
                convert_to_json_file(args.input_dir, f, quiet=args.quiet)
        case OutputFormat.CSV:
            with open(args.output_file + '.csv', 'w', newline='') as f:
                convert_to_csv_file(args.input_dir, f, quiet=args.quiet)


if __name__ == '__main__':
    main()
